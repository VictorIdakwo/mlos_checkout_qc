import streamlit as st
import sqlite3
import pandas as pd
import tempfile
import os
import re
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MLOS Checkout QC",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── WHITE THEME CSS ────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }

  /* Force white background */
  .stApp { background-color: #ffffff !important; color: #1e293b !important; }
  section[data-testid="stSidebar"] { background-color: #f8fafc !important; border-right: 1px solid #e2e8f0; }
  [data-testid="stSidebar"] * { color: #1e293b !important; }

  /* Header */
  .app-header {
    background: #1d4ed8;
    border-radius: 10px; padding: 22px 28px; margin-bottom: 22px;
  }
  .app-header h1 { font-size: 1.8rem; font-weight: 700; color: #ffffff; margin: 0 0 4px 0; }
  .app-header p  { font-size: 0.88rem; color: #bfdbfe; margin: 0; }

  /* Verdict banners */
  .banner-pass {
    background: #f0fdf4; border: 1px solid #86efac;
    border-radius: 8px; padding: 14px 20px; margin-bottom: 16px;
    color: #15803d; font-weight: 600; font-size: 0.95rem;
  }
  .banner-fail {
    background: #fff7ed; border: 1px solid #fdba74;
    border-radius: 8px; padding: 14px 20px; margin-bottom: 16px;
    color: #c2410c; font-weight: 600; font-size: 0.95rem;
  }

  /* Section titles */
  .sec-title {
    font-size: 1rem; font-weight: 700; color: #1e40af;
    border-bottom: 2px solid #dbeafe; padding-bottom: 8px; margin-bottom: 14px;
  }

  /* Report verdict */
  .report-verdict-pass {
    background: #f0fdf4; border: 2px solid #86efac;
    border-radius: 10px; padding: 22px; text-align: center;
    color: #15803d; font-size: 1.25rem; font-weight: 700; margin-bottom: 20px;
  }
  .report-verdict-fail {
    background: #fff1f2; border: 2px solid #fca5a5;
    border-radius: 10px; padding: 22px; text-align: center;
    color: #be123c; font-size: 1.25rem; font-weight: 700; margin-bottom: 20px;
  }

  /* Report metadata rows */
  .report-card {
    background: #f8fafc; border: 1px solid #e2e8f0;
    border-radius: 10px; padding: 20px 24px; margin-bottom: 20px;
  }
  .report-row {
    display: flex; justify-content: space-between; align-items: center;
    padding: 7px 0; border-bottom: 1px solid #f1f5f9; font-size: 0.88rem;
  }
  .report-row .lbl { color: #64748b; }
  .report-row .val { color: #1e293b; font-weight: 600; }
  .report-row .val-pass { color: #15803d; font-weight: 700; }
  .report-row .val-fail { color: #dc2626; font-weight: 700; }
  .report-row .val-div { color: #e2e8f0; }
</style>
""", unsafe_allow_html=True)

# ─── Constants ───────────────────────────────────────────────────────────────────
MLOS_VIEW    = "master_list_settlement_update_view"
TAKEOFF_VIEW = "mlos_takeoffpoint_view"

VALID_DAY   = {"1","1_2","1_2_3","1_2_3_4","2","2_3","2_3_4","3","3_4","4","NA"}
VALID_YN_NA = {"Y","N","NA"}
VALID_YN    = {"Y","N"}
VALID_ACC   = {"Fully Accessible","Partially Accessible","Inaccessible"}
VALID_HAB   = {"Abandoned","Migrated","Inhabited","Partially Inhabited"}
UUID_RE     = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')
EDITOR_RE   = re.compile(r'^[a-z]+\.[a-z]+$')
NULLABLE    = {"primarysettlement_name","alternate_name","reasons_for_inaccessibility"}
YN_NA_COLS  = {"highrisk","slums","densely_populated","hard2reach","border",
               "normadic","riverine","fulani","team_code"}

# ─── Helpers ─────────────────────────────────────────────────────────────────────
def load_sqlite(uploaded_file):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite")
    tmp.write(uploaded_file.read())
    tmp.close()
    return sqlite3.connect(tmp.name), tmp.name

def load_view(conn, view_name):
    cur = conn.cursor()
    cur.execute(f'PRAGMA table_info("{view_name}")')
    cols_info = cur.fetchall()
    safe = [c[1] for c in cols_info
            if c[2] is None or not any(k in str(c[2]).upper()
                                       for k in ("BLOB","POINT","GEOMETRY"))]
    col_str = ", ".join(f'"{c}"' for c in safe)
    return pd.read_sql_query(f'SELECT {col_str} FROM "{view_name}"', conn)

def load_csv(uploaded_file):
    """Load a CSV file as the MLoS table. Returns (mlos_df, empty takeoff_df)."""
    uploaded_file.seek(0)
    mlos_df = pd.read_csv(uploaded_file, dtype=str)
    takeoff_df = pd.DataFrame()
    return mlos_df, takeoff_df

def load_xlsx(uploaded_file):
    """Load an XLSX/XLS file. Expects two sheets: mlos (sheet 1) and takeoffpoint (sheet 2).
    Sheet names may match the view names or simplified aliases."""
    uploaded_file.seek(0)
    xl = pd.ExcelFile(uploaded_file)
    sheet_names = xl.sheet_names

    # Resolve mlos sheet — prefer view name, then 'mlos', then first sheet
    mlos_aliases = {MLOS_VIEW.lower(), "mlos", "master_list", "settlements"}
    mlos_sheet = next(
        (s for s in sheet_names if s.lower() in mlos_aliases),
        sheet_names[0]
    )

    # Resolve takeoff sheet — prefer view name, then 'takeoffpoint', then second sheet
    takeoff_aliases = {TAKEOFF_VIEW.lower(), "takeoffpoint", "takeoff", "takeoff_point"}
    takeoff_sheet = next(
        (s for s in sheet_names if s.lower() in takeoff_aliases and s != mlos_sheet),
        sheet_names[1] if len(sheet_names) > 1 else None
    )

    mlos_df    = xl.parse(mlos_sheet, dtype=str)
    takeoff_df = xl.parse(takeoff_sheet, dtype=str) if takeoff_sheet else pd.DataFrame()
    return mlos_df, takeoff_df

def get_uploaded_data(uploaded_file, progress=None):
    cache_key = (uploaded_file.name, getattr(uploaded_file, "size", None))
    if st.session_state.get("uploaded_cache_key") == cache_key:
        if progress is not None:
            progress.progress(100)
        return st.session_state["mlos_df"], st.session_state["takeoff_df"]

    if progress is not None:
        progress.progress(10)

    ext = uploaded_file.name.rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xls"):
        if progress is not None:
            progress.progress(35)
        mlos_df, takeoff_df = load_xlsx(uploaded_file)
        if progress is not None:
            progress.progress(90)
    elif ext == "csv":
        if progress is not None:
            progress.progress(35)
        mlos_df, takeoff_df = load_csv(uploaded_file)
        if progress is not None:
            progress.progress(90)
    else:
        conn, tmp_path = load_sqlite(uploaded_file)
        try:
            if progress is not None:
                progress.progress(35)
            mlos_df    = load_view(conn, MLOS_VIEW)
            if progress is not None:
                progress.progress(65)
            takeoff_df = load_view(conn, TAKEOFF_VIEW)
            if progress is not None:
                progress.progress(90)
        finally:
            conn.close()
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    st.session_state["uploaded_cache_key"] = cache_key
    st.session_state["mlos_df"] = mlos_df
    st.session_state["takeoff_df"] = takeoff_df
    if progress is not None:
        progress.progress(100)
    return mlos_df, takeoff_df


def pct(v, t): return f"{v/t*100:.1f}%" if t else "0%"

# Vectorized regex checks — avoid per-row Python function calls
def vec_is_uuid(series: "pd.Series") -> "pd.Series":
    s = series.astype(str).str.strip()
    return s.str.match(UUID_RE.pattern, na=False)

def vec_is_editor(series: "pd.Series") -> "pd.Series":
    s = series.astype(str).str.strip()
    return s.str.match(EDITOR_RE.pattern, na=False)

# ─── Schema QC ───────────────────────────────────────────────────────────────────
MLOS_REQUIRED_COLS = {
    "state_code", "state_name", "lga_code", "lga_name", "ward_name", "ward_code",
    "takeoffpoint", "takeoffpoint_code", "settlement_name", "primarysettlement_name",
    "alternate_name", "latitude", "longitude", "security_compromised",
    "accessibility_status", "reasons_for_inaccessibility", "habitational_status",
    "set_population", "set_target", "number_of_houses", "noncompliant_household",
    "team_code", "day_of_activity", "urban", "rural", "highrisk", "slums",
    "densely_populated", "hard2reach", "border", "normadic", "scattered",
    "riverine", "fulani", "timestamp", "source", "last_updated", "editor",
    "globalid", "fc_globalid", "settlementarea_globalid",
}
TAKEOFF_REQUIRED_COLS = {"name", "code", "wardcode", "globalid"}

def run_schema_qc(mlos: pd.DataFrame, takeoff: pd.DataFrame):
    """
    Schema alignment QC — checks required columns are present.
    Returns check entries (same format as run_mlos_qc) and a detail DataFrame.
    Does NOT stop the QC process.
    """
    checks, details = [], []

    mlos_cols    = set(mlos.columns.str.strip().str.lower())
    missing_mlos = sorted(MLOS_REQUIRED_COLS - mlos_cols)
    n_mlos_total = len(MLOS_REQUIRED_COLS)
    n_mlos_miss  = len(missing_mlos)
    checks.append({
        "Rule#": "S1", "QC Check": "MLoS Schema Alignment",
        "Description": f"{n_mlos_total - n_mlos_miss}/{n_mlos_total} required MLoS columns present",
        "Failing Rows": n_mlos_miss, "Total Rows": n_mlos_total,
        "Fail %": pct(n_mlos_miss, n_mlos_total),
        "Status": "❌ FAIL" if n_mlos_miss else "✅ PASS",
    })
    if missing_mlos:
        details.append(pd.DataFrame([{
            "Rule#": "S1", "Rule": "MLoS Schema Alignment",
            "Table": "MLoS", "Missing Column": col,
            "Impact": "QC rule(s) referencing this column will be skipped or may return incorrect results",
        } for col in missing_mlos]))

    if not takeoff.empty and len(takeoff.columns) > 0:
        tp_cols    = set(takeoff.columns.str.strip().str.lower())
        missing_tp = sorted(TAKEOFF_REQUIRED_COLS - tp_cols)
        n_tp_total = len(TAKEOFF_REQUIRED_COLS)
        n_tp_miss  = len(missing_tp)
        checks.append({
            "Rule#": "S2", "QC Check": "Takeoffpoint Schema Alignment",
            "Description": f"{n_tp_total - n_tp_miss}/{n_tp_total} required Takeoffpoint columns present",
            "Failing Rows": n_tp_miss, "Total Rows": n_tp_total,
            "Fail %": pct(n_tp_miss, n_tp_total),
            "Status": "❌ FAIL" if n_tp_miss else "✅ PASS",
        })
        if missing_tp:
            details.append(pd.DataFrame([{
                "Rule#": "S2", "Rule": "Takeoffpoint Schema Alignment",
                "Table": "Takeoffpoint", "Missing Column": col,
                "Impact": "QC rule(s) referencing this column will be skipped or may return incorrect results",
            } for col in missing_tp]))

    schema_detail = pd.concat(details, ignore_index=True) if details else pd.DataFrame()
    return checks, schema_detail

def build_schema_report_xlsx(schema_detail: pd.DataFrame) -> bytes:
    """Produce a downloadable Excel report listing schema mismatches."""
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        schema_detail.to_excel(writer, index=False, sheet_name="Schema Issues")
        ws = writer.sheets["Schema Issues"]
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in col_cells) + 4
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length, 60)
    return out.getvalue()

# ─── Longitudinal MLoS Report ────────────────────────────────────────────────────
def make_longitudinal_df(mlos_df: pd.DataFrame,
                          mlos_checks: list,
                          mlos_detail: pd.DataFrame) -> pd.DataFrame:
    """
    Build a wide/longitudinal DataFrame where:
    - Each row is a unique settlement (appears once only)
    - Each QC rule adds a boolean column: True = error, False = no error
    - Only rows with at least one True are returned
    """
    result = mlos_df.copy()
    result["_orig_idx"] = mlos_df.index

    rule_cols = []
    for check in mlos_checks:
        rn       = check["Rule#"]
        col_name = f"Rule_{rn} | {check['QC Check']}"
        rule_cols.append(col_name)
        if "FAIL" in check["Status"] and not mlos_detail.empty:
            failing_idx = set(mlos_detail.loc[mlos_detail["Rule#"] == rn].index)
            result[col_name] = result["_orig_idx"].isin(failing_idx)
        else:
            result[col_name] = False

    result = result.drop(columns=["_orig_idx"])
    error_mask = result[rule_cols].any(axis=1)
    return result[error_mask].reset_index(drop=True)


def build_longitudinal_mlos(mlos_df: pd.DataFrame,
                             mlos_checks: list,
                             mlos_detail: pd.DataFrame) -> bytes:
    """Write the longitudinal DataFrame to a styled Excel file and return bytes."""
    error_df = make_longitudinal_df(mlos_df, mlos_checks, mlos_detail)

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        error_df.to_excel(writer, index=False, sheet_name="MLoS Issues — Longitudinal")
        ws          = writer.sheets["MLoS Issues — Longitudinal"]
        hdr_fill    = PatternFill("solid", fgColor="1D4ED8")
        err_fill    = PatternFill("solid", fgColor="FEE2E2")
        ok_fill     = PatternFill("solid", fgColor="F0FDF4")
        hdr_font    = Font(bold=True, color="FFFFFF")
        body_font   = Font(size=10)
        center      = Alignment(horizontal="center")

        rule_col_indices = {
            i + 1 for i, h in enumerate(error_df.columns)
            if str(h).startswith("Rule_")
        }

        for ci, cell in enumerate(ws[1], 1):
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center

        for ri in range(2, len(error_df) + 2):
            for ci in range(1, len(error_df.columns) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.font = body_font
                if ci in rule_col_indices:
                    cell.alignment = center
                    cell.fill = err_fill if cell.value else ok_fill

        for ci, col_name in enumerate(error_df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = (
                36 if str(col_name).startswith("Rule_") else 18
            )

    return out.getvalue()

# ─── Boundary Reference Loading ──────────────────────────────────────────────────
BOUNDARY_REF_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ward_boundary_ref.csv")
BOUNDARY_BBOX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ward_boundary_bbox.csv")

@st.cache_data(show_spinner=False)
def load_boundary_refs():
    """Load ward code reference and bounding box lookup tables."""
    ref_df  = pd.read_csv(BOUNDARY_REF_PATH,  dtype=str) if os.path.exists(BOUNDARY_REF_PATH)  else pd.DataFrame()
    bbox_df = pd.read_csv(BOUNDARY_BBOX_PATH, dtype=str) if os.path.exists(BOUNDARY_BBOX_PATH) else pd.DataFrame()
    for col in ["min_lon","min_lat","max_lon","max_lat"]:
        if col in bbox_df.columns:
            bbox_df[col] = pd.to_numeric(bbox_df[col], errors="coerce")
    return ref_df, bbox_df

# ─── QC Engine — Ward Boundary ───────────────────────────────────────────────────
def run_ward_boundary_qc(mlos: pd.DataFrame, ref_df: pd.DataFrame, bbox_df: pd.DataFrame):
    checks, details = [], []
    if "ward_code" not in mlos.columns:
        return checks, pd.DataFrame()

    total  = len(mlos)
    id_col = "ogc_fid" if "ogc_fid" in mlos.columns else mlos.columns[0]
    BASE   = [c for c in [id_col,"state_name","lga_name","ward_code","settlement_name"] if c in mlos.columns]

    def add(num, rule, desc, mask, extra=None):
        n    = int(mask.sum())
        cols = BASE + [c for c in (extra or []) if c in mlos.columns and c not in BASE]
        checks.append({"Rule#": num, "QC Check": rule, "Description": desc,
                        "Failing Rows": n, "Total Rows": total,
                        "Fail %": pct(n, total),
                        "Status": "❌ FAIL" if n else "✅ PASS"})
        if n:
            sub = mlos[mask][cols].copy()
            sub.insert(0, "Rule#", num)
            sub.insert(1, "Rule", rule)
            details.append(sub)

    # ── Pre-filter boundary reference to only the state_codes and lga_codes
    # present in the uploaded file — shrinks the search space significantly
    filtered_ref  = ref_df.copy()  if not ref_df.empty  else ref_df
    filtered_bbox = bbox_df.copy() if not bbox_df.empty else bbox_df

    if not ref_df.empty:
        if "state_code" in mlos.columns and "state_code" in ref_df.columns:
            mlos_states = set(mlos["state_code"].dropna().astype(str).str.strip())
            filtered_ref  = filtered_ref[filtered_ref["state_code"].astype(str).str.strip().isin(mlos_states)]
            filtered_bbox = filtered_bbox[
                filtered_bbox["ward_code"].isin(set(filtered_ref["ward_code"].dropna()))
            ] if not filtered_bbox.empty and "ward_code" in filtered_bbox.columns else filtered_bbox

        if "lga_code" in mlos.columns and "lga_code" in ref_df.columns:
            mlos_lgas = set(mlos["lga_code"].dropna().astype(str).str.strip())
            filtered_ref  = filtered_ref[filtered_ref["lga_code"].astype(str).str.strip().isin(mlos_lgas)]
            filtered_bbox = filtered_bbox[
                filtered_bbox["ward_code"].isin(set(filtered_ref["ward_code"].dropna()))
            ] if not filtered_bbox.empty and "ward_code" in filtered_bbox.columns else filtered_bbox

    # B1 — ward_code must exist in the (pre-filtered) boundary reference
    if not filtered_ref.empty and "ward_code" in filtered_ref.columns:
        valid_wards = set(filtered_ref["ward_code"].dropna().str.strip())
        add("B1", "Ward Code — Boundary Reference",
            "ward_code must exist in the admin ward boundary reference for the file's state(s) and LGA(s)",
            ~mlos["ward_code"].astype(str).str.strip().isin(valid_wards),
            ["ward_code"])

    # B2 — lat/lon must fall within the bounding box of the declared ward_code
    if (not filtered_bbox.empty and "ward_code" in filtered_bbox.columns
            and "latitude" in mlos.columns and "longitude" in mlos.columns):
        # Vectorised merge against the pre-filtered bounding boxes only
        chk = pd.DataFrame({
            "ward_code": mlos["ward_code"].astype(str).str.strip().values,
            "lat":       pd.to_numeric(mlos["latitude"],  errors="coerce").values,
            "lon":       pd.to_numeric(mlos["longitude"], errors="coerce").values,
        }, index=mlos.index)

        bbox_ref = filtered_bbox[["ward_code","min_lon","min_lat","max_lon","max_lat"]].copy()
        for col in ["min_lon","min_lat","max_lon","max_lat"]:
            bbox_ref[col] = pd.to_numeric(bbox_ref[col], errors="coerce")

        merged = chk.merge(bbox_ref, on="ward_code", how="left")
        merged.index = mlos.index

        has_coords   = merged["lat"].notna() & merged["lon"].notna()
        in_reference = merged["min_lon"].notna()
        out_of_box   = (
            (merged["lon"] < merged["min_lon"]) |
            (merged["lon"] > merged["max_lon"]) |
            (merged["lat"] < merged["min_lat"]) |
            (merged["lat"] > merged["max_lat"])
        )
        mask_b2 = has_coords & in_reference & out_of_box
        add("B2", "Coordinates — Within Ward Boundary",
            "latitude/longitude must fall within the bounding box of the declared ward_code",
            mask_b2, ["ward_code","latitude","longitude"])

    detail_df = pd.concat(details, ignore_index=True) if details else pd.DataFrame()
    return checks, detail_df

# ─── Email Sender ─────────────────────────────────────────────────────────────────
def send_qc_email(filename: str, all_checks: list, mlos_fail_rows: int,
                  tp_fail_rows: int, schema_detail: pd.DataFrame,
                  boundary_fail_rows: int):
    """Send QC summary email using SMTP credentials stored in st.secrets."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    TO  = ["adanna.alex@ehealthnigeria.org"]
    CC  = ["fashoto.busayo@ehealthnigeria.org",
           "victor.idakwo@ehealthnigeria.org",
           "oluwadamilare.akindipe@ehealthnigeria.org"]

    n_fail  = sum(1 for c in all_checks if "FAIL" in c["Status"])
    n_pass  = len(all_checks) - n_fail
    verdict = "CLEAN ✅" if n_fail == 0 else f"FAILING ❌ ({n_fail} check(s))"

    schema_lines = ""
    if not schema_detail.empty:
        missing = schema_detail[["Table","Missing Column"]].to_string(index=False)
        schema_lines = f"\nMissing Columns:\n{missing}\n"

    body = f"""MLoS QC Report — {filename}
{"="*60}
Overall Verdict   : {verdict}
Checks Run        : {len(all_checks)}
Checks Passing    : {n_pass}
Checks Failing    : {n_fail}

MLoS Issue Rows       : {mlos_fail_rows:,}
Takeoffpoint Issue Rows: {tp_fail_rows:,}
Boundary Issue Rows   : {boundary_fail_rows:,}
{schema_lines}
Check-by-Check Summary:
{"-"*60}
"""
    for c in all_checks:
        body += f"[{c['Status']}] Rule {c['Rule#']}: {c['QC Check']} — {c['Failing Rows']} failing / {c['Total Rows']} ({c['Fail %']})\n"

    body += f"\n{'='*60}\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"

    msg = MIMEMultipart()
    msg["From"]    = st.secrets.get("smtp_user", "")
    msg["To"]      = ", ".join(TO)
    msg["Cc"]      = ", ".join(CC)
    msg["Subject"] = f"MLoS QC checks for {filename}"
    msg.attach(MIMEText(body, "plain"))

    smtp_host = st.secrets.get("smtp_host", "smtp.gmail.com")
    smtp_port = int(st.secrets.get("smtp_port", 587))
    smtp_user = st.secrets.get("smtp_user", "")
    smtp_pass = st.secrets.get("smtp_pass", "")

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, TO + CC, msg.as_string())

# ─── Auto Correction ─────────────────────────────────────────────────────────────
FLAG_COLS = [
    "highrisk", "slums", "densely_populated", "hard2reach",
    "border", "normadic", "scattered", "riverine", "fulani",
]

def auto_correct_mlos(mlos: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    """
    Apply three auto-corrections to the MLoS DataFrame.
    Returns a corrected copy and a log of changes made.
    """
    import uuid as _uuid
    df  = mlos.copy()
    log = []   # list of {Column, Rule, Rows Fixed}

    # ── Correction 1: Flag columns NULL → "NA" ───────────────────────────────
    for col in FLAG_COLS:
        if col not in df.columns:
            continue
        mask = df[col].isna() | df[col].astype(str).str.strip().eq("")
        n = int(mask.sum())
        if n:
            df.loc[mask, col] = "NA"
            log.append({"Column": col,
                        "Correction": "NULL → NA",
                        "Rows Fixed": n})

    # ── Correction 2: Fully Accessible + NULL reason → "NA" ──────────────────
    if "accessibility_status" in df.columns and "reasons_for_inaccessibility" in df.columns:
        mask = (
            df["accessibility_status"].astype(str).str.strip().eq("Fully Accessible") &
            (df["reasons_for_inaccessibility"].isna() |
             df["reasons_for_inaccessibility"].astype(str).str.strip().eq(""))
        )
        n = int(mask.sum())
        if n:
            df.loc[mask, "reasons_for_inaccessibility"] = "NA"
            log.append({"Column": "reasons_for_inaccessibility",
                        "Correction": "NULL → NA (Fully Accessible rows)",
                        "Rows Fixed": n})

    # ── Correction 3: Invalid globalid → strip braces / regenerate UUID ──────
    if "globalid" in df.columns:
        # Step 1: strip leading { and trailing } if present
        brace_mask = df["globalid"].astype(str).str.match(r"^\{.*\}$", na=False)
        n_brace = int(brace_mask.sum())
        if n_brace:
            df.loc[brace_mask, "globalid"] = (
                df.loc[brace_mask, "globalid"]
                  .astype(str).str.strip().str.strip("{}")
            )

        # Step 2: regenerate UUID for any still-invalid globalid
        invalid_mask = ~vec_is_uuid(df["globalid"])
        n_gen = int(invalid_mask.sum())
        if n_gen:
            df.loc[invalid_mask, "globalid"] = [
                str(_uuid.uuid4()) for _ in range(n_gen)
            ]

        total_fixed = n_brace + n_gen
        if total_fixed:
            log.append({"Column": "globalid",
                        "Correction": f"Braces stripped ({n_brace} rows); UUID regenerated ({n_gen} rows)",
                        "Rows Fixed": total_fixed})

    return df, log


def build_corrected_excel(df: pd.DataFrame) -> bytes:
    """Write the corrected MLoS DataFrame to an Excel file and return bytes."""
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="MLoS Corrected")
        ws       = writer.sheets["MLoS Corrected"]
        hdr_fill = PatternFill("solid", fgColor="1D4ED8")
        hdr_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
        for ci, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(
                len(str(col)) + 4, 14)
    return out.getvalue()


# ─── QC Engine — MLoS ────────────────────────────────────────────────────────────
def run_mlos_qc(mlos: pd.DataFrame, takeoff: pd.DataFrame):
    checks, details = [], []
    total  = len(mlos)
    id_col = "ogc_fid" if "ogc_fid" in mlos.columns else mlos.columns[0]
    BASE   = [c for c in [id_col,"state_name","lga_name","ward_code",
                           "settlement_name","takeoffpoint","takeoffpoint_code"]
              if c in mlos.columns]

    def add(num, rule, desc, mask, extra=None):
        n    = int(mask.sum())
        cols = BASE + [c for c in (extra or []) if c in mlos.columns and c not in BASE]
        checks.append({"Rule#": num, "QC Check": rule, "Description": desc,
                       "Failing Rows": n, "Total Rows": total,
                       "Fail %": pct(n, total),
                       "Status": "❌ FAIL" if n else "✅ PASS"})
        if n:
            sub = mlos[mask][cols].copy()
            sub.insert(0, "Rule#", num)
            sub.insert(1, "Rule", rule)
            details.append(sub)

    # R2
    if "takeoffpoint" in mlos.columns and "name" in takeoff.columns:
        valid = set(takeoff["name"].dropna().str.strip())
        add("2","Takeoffpoint Name Match",
            "takeoffpoint in MLoS must match name in Takeoffpoint table",
            ~mlos["takeoffpoint"].astype(str).str.strip().isin(valid), ["takeoffpoint"])

    # R3
    if "takeoffpoint_code" in mlos.columns and "code" in takeoff.columns:
        valid = set(takeoff["code"].dropna().str.strip())
        add("3","Takeoffpoint Code Match",
            "takeoffpoint_code in MLoS must match code in Takeoffpoint table",
            ~mlos["takeoffpoint_code"].astype(str).str.strip().isin(valid), ["takeoffpoint_code"])

    # R4
    if "ward_code" in mlos.columns and "wardcode" in takeoff.columns:
        valid = set(takeoff["wardcode"].dropna().str.strip())
        add("4","Ward Code Match",
            "ward_code in MLoS must match wardcode in Takeoffpoint table",
            ~mlos["ward_code"].astype(str).str.strip().isin(valid), ["ward_code"])

    # R5 – required fields not null
    non_null = [c for c in mlos.columns
                if c not in NULLABLE and
                   c not in {"ROWID","ogc_fid","geom","fc_globalid","settlementarea_globalid"}]
    null_mask = (mlos[non_null].isnull() | mlos[non_null].astype(str).eq("")).any(axis=1)
    add("5","No Null in Required Fields",
        "All fields except primarysettlement_name, alternate_name, reasons_for_inaccessibility must not be null",
        null_mask)

    # R6
    if "security_compromised" in mlos.columns:
        add("6","Security Compromised Y/N",
            "security_compromised must be Y or N",
            ~mlos["security_compromised"].astype(str).str.strip().isin({"Y","N"}),
            ["security_compromised"])

    # R7
    if "accessibility_status" in mlos.columns:
        add("7","Accessibility Status Valid",
            "accessibility_status must be: Fully Accessible, Partially Accessible, or Inaccessible",
            ~mlos["accessibility_status"].astype(str).str.strip().isin(VALID_ACC),
            ["accessibility_status"])

    # R8
    if "accessibility_status" in mlos.columns and "reasons_for_inaccessibility" in mlos.columns:
        needs     = mlos["accessibility_status"].isin(["Partially Accessible","Inaccessible"])
        no_reason = (mlos["reasons_for_inaccessibility"].isna() |
                     mlos["reasons_for_inaccessibility"].astype(str).str.strip().eq(""))
        add("8","Reason for Inaccessibility Required",
            "Partially Accessible & Inaccessible settlements must have reasons_for_inaccessibility",
            needs & no_reason, ["accessibility_status","reasons_for_inaccessibility"])

    # R9
    if "habitational_status" in mlos.columns:
        add("9","Habitational Status Valid",
            "habitational_status must be: Abandoned, Migrated, Inhabited, or Partially Inhabited",
            ~mlos["habitational_status"].astype(str).str.strip().isin(VALID_HAB),
            ["habitational_status"])

    # R10
    for col in ["set_target","number_of_houses"]:
        if col in mlos.columns and "set_population" in mlos.columns:
            mask = (pd.to_numeric(mlos[col], errors="coerce") >
                    pd.to_numeric(mlos["set_population"], errors="coerce")).fillna(False)
            add("10",f"{col} ≤ set_population",
                f"{col} must not be higher than set_population",
                mask, [col,"set_population"])

    # R12
    if "day_of_activity" in mlos.columns:
        add("12","Day of Activity Valid",
            "day_of_activity must be one of: 1, 1_2, 1_2_3, 1_2_3_4, 2, 2_3, 2_3_4, 3, 3_4, 4, NA",
            ~mlos["day_of_activity"].astype(str).str.strip().isin(VALID_DAY),
            ["day_of_activity"])

    # R13
    for col in ["urban","rural","scattered"]:
        if col in mlos.columns:
            add("13",f"{col} is Y or N",
                f"{col} must be Y or N and must not be null",
                ~mlos[col].astype(str).str.strip().isin(VALID_YN), [col])

    if "urban" in mlos.columns and "rural" in mlos.columns:
        mask = ((mlos["urban"].astype(str).str.strip()=="Y") &
                (mlos["rural"].astype(str).str.strip()=="Y"))
        add("13a","Cannot be both Urban and Rural",
            "A settlement cannot be both urban=Y and rural=Y", mask, ["urban","rural"])

    if "urban" in mlos.columns and "scattered" in mlos.columns:
        mask = ((mlos["urban"].astype(str).str.strip()=="Y") &
                (mlos["scattered"].astype(str).str.strip()=="Y"))
        add("13b","Urban cannot be Scattered",
            "Urban settlements cannot be categorised as scattered", mask, ["urban","scattered"])

    # R14
    for col in YN_NA_COLS:
        if col in mlos.columns:
            add("14",f"{col} = Y/N/NA",
                f"{col} must be Y, N, or NA",
                ~mlos[col].astype(str).str.strip().isin(VALID_YN_NA), [col])

    # R15
    if "source" in mlos.columns:
        add("15","Source = MLoS","source field must start with 'MLoS'",
            ~mlos["source"].astype(str).str.strip().str.startswith("MLoS"), ["source"])

    # R16
    if "editor" in mlos.columns:
        add("16","Editor Format (firstname.surname)",
            "editor must be in format: firstname.surname (all lowercase)",
            ~vec_is_editor(mlos["editor"]), ["editor"])

    # R17
    if "globalid" in mlos.columns:
        add("17","globalid is UUID",
            "globalid must be a valid UUID (xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)",
            ~vec_is_uuid(mlos["globalid"]), ["globalid"])

    # Keep original mlos_df indices so make_longitudinal_df can match failing rows correctly
    return checks, pd.concat(details) if details else pd.DataFrame()


# ─── QC Engine — Takeoffpoint ────────────────────────────────────────────────────
def run_takeoff_qc(takeoff: pd.DataFrame, mlos: pd.DataFrame):
    if takeoff.empty or len(takeoff.columns) == 0:
        return [], pd.DataFrame()
    checks, details = [], []
    total  = len(takeoff)
    id_col = "ogc_fid" if "ogc_fid" in takeoff.columns else takeoff.columns[0]
    BASE   = [c for c in [id_col,"name","wardcode","code"] if c in takeoff.columns]

    def add(num, rule, desc, mask, extra=None):
        n    = int(mask.sum())
        cols = BASE + [c for c in (extra or []) if c in takeoff.columns and c not in BASE]
        checks.append({"Rule#": num, "QC Check": rule, "Description": desc,
                       "Failing Rows": n, "Total Rows": total,
                       "Fail %": pct(n, total),
                       "Status": "❌ FAIL" if n else "✅ PASS"})
        if n:
            sub = takeoff[mask][cols].copy()
            sub.insert(0, "Rule#", num)
            sub.insert(1, "Rule", rule)
            details.append(sub)

    if "name" in takeoff.columns and "takeoffpoint" in mlos.columns:
        add("TP2","Name matches MLoS takeoffpoint",
            "name in Takeoffpoint must match takeoffpoint column in MLoS",
            ~takeoff["name"].astype(str).str.strip().isin(set(mlos["takeoffpoint"].dropna().str.strip())),
            ["name"])

    if "code" in takeoff.columns and "takeoffpoint_code" in mlos.columns:
        add("TP3","Code matches MLoS takeoffpoint_code",
            "code in Takeoffpoint must match takeoffpoint_code in MLoS",
            ~takeoff["code"].astype(str).str.strip().isin(set(mlos["takeoffpoint_code"].dropna().str.strip())),
            ["code"])

    if "wardcode" in takeoff.columns and "ward_code" in mlos.columns:
        add("TP4","wardcode matches MLoS ward_code",
            "wardcode in Takeoffpoint must match ward_code in MLoS",
            ~takeoff["wardcode"].astype(str).str.strip().isin(set(mlos["ward_code"].dropna().str.strip())),
            ["wardcode"])

    if "globalid" in takeoff.columns:
        add("TP5","globalid is UUID","globalid must be a valid UUID",
            ~vec_is_uuid(takeoff["globalid"]), ["globalid"])

    return checks, pd.concat(details, ignore_index=True) if details else pd.DataFrame()


# ─── Excel Report Builder ─────────────────────────────────────────────────────────
def build_excel_report(filename, mlos_checks, mlos_detail, tp_checks, tp_detail,
                        mlos_df, takeoff_df):
    out = BytesIO()
    wb  = openpyxl.Workbook()

    BLU   = "1D4ED8"
    GRN   = "15803D"; LGN = "F0FDF4"
    RED   = "DC2626"; LRD = "FFF1F2"
    LGY   = "F8FAFC"
    WHT   = "FFFFFF"

    hdr_fill  = PatternFill("solid", fgColor=BLU)
    hdr_font  = Font(bold=True, color=WHT, size=10, name="Calibri")
    pass_fill = PatternFill("solid", fgColor=LGN)
    fail_fill = PatternFill("solid", fgColor=LRD)
    lgry_fill = PatternFill("solid", fgColor=LGY)
    body_font = Font(size=10, name="Calibri")
    bold_font = Font(bold=True, size=10, name="Calibri")
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="D1D5DB")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_row(ws, row, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=col)
            c.fill=hdr_fill; c.font=hdr_font; c.alignment=ctr; c.border=bdr

    total_fail = (sum(1 for c in mlos_checks if "FAIL" in c["Status"]) +
                  sum(1 for c in tp_checks   if "FAIL" in c["Status"]))

    # ── Sheet 1: Cover ──
    ws = wb.active; ws.title = "QC Report Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 52
    ws.merge_cells("A1:B1")
    ws["A1"].value = "MLOS CHECKOUT — QUALITY CONTROL REPORT"
    ws["A1"].font  = Font(bold=True, size=15, color=WHT, name="Calibri")
    ws["A1"].fill  = PatternFill("solid", fgColor=BLU)
    ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 38

    meta = [
        ("File",                   filename),
        ("Generated At",           datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("MLoS Rows",              len(mlos_df)),
        ("Takeoffpoint Rows",      len(takeoff_df)),
        ("MLoS Checks Run",        len(mlos_checks)),
        ("MLoS Checks Failing",    sum(1 for c in mlos_checks if "FAIL" in c["Status"])),
        ("TP Checks Run",          len(tp_checks)),
        ("TP Checks Failing",      sum(1 for c in tp_checks   if "FAIL" in c["Status"])),
        ("Total MLoS Issue Rows",  len(mlos_detail)),
        ("Total TP Issue Rows",    len(tp_detail)),
    ]
    for ri, (k, v) in enumerate(meta, 3):
        ka = ws.cell(row=ri, column=1, value=k); ka.font=bold_font; ka.border=bdr
        vb = ws.cell(row=ri, column=2, value=v); vb.font=body_font; vb.border=bdr; vb.alignment=lft
        if ri % 2 == 0: ka.fill=lgry_fill; vb.fill=lgry_fill

    ws.merge_cells("A14:B14")
    vd = ws["A14"]
    vd.value     = "✅  ALL CHECKS PASSED — File is CLEAN" if total_fail == 0 \
                   else f"❌  {total_fail} CHECK(S) FAILING — Review issue sheets"
    vd.fill      = PatternFill("solid", fgColor=GRN if total_fail==0 else RED)
    vd.font      = Font(bold=True, size=12, color=WHT, name="Calibri")
    vd.alignment = ctr
    ws.row_dimensions[14].height = 30

    # helper to write a QC summary sheet
    def write_summary(ws_s, title, checks_list):
        ws_s.sheet_view.showGridLines = False
        ws_s.merge_cells("A1:G1")
        ws_s["A1"].value     = title
        ws_s["A1"].font      = Font(bold=True, size=13, color=WHT, name="Calibri")
        ws_s["A1"].fill      = PatternFill("solid", fgColor=BLU)
        ws_s["A1"].alignment = ctr
        ws_s.row_dimensions[1].height = 30
        cols = ["Status","Rule#","QC Check","Description","Failing Rows","Total Rows","Fail %"]
        hdr_row(ws_s, 2, cols)
        for ri, row in enumerate(checks_list, 3):
            is_fail = "FAIL" in row["Status"]
            for ci, key in enumerate(cols, 1):
                c = ws_s.cell(row=ri, column=ci, value=row[key])
                c.font=body_font; c.border=bdr
                c.alignment = ctr if ci in (1,2,5,6,7) else lft
                c.fill = fail_fill if is_fail else pass_fill
        for i, w in enumerate([14,8,32,65,14,12,9], 1):
            ws_s.column_dimensions[get_column_letter(i)].width = w

    # helper to write issue detail sheet
    def write_detail(ws_d, title, detail_df):
        ws_d.sheet_view.showGridLines = False
        if detail_df.empty:
            ws_d["A1"].value     = f"✅ No issues — {title} passed all checks"
            ws_d["A1"].font      = Font(bold=True, size=12, color=WHT, name="Calibri")
            ws_d["A1"].fill      = PatternFill("solid", fgColor=GRN)
            ws_d["A1"].alignment = ctr
            return
        nc = len(detail_df.columns)
        ws_d.merge_cells(f"A1:{get_column_letter(nc)}1")
        ws_d["A1"].value     = title
        ws_d["A1"].font      = Font(bold=True, size=12, color=WHT, name="Calibri")
        ws_d["A1"].fill      = PatternFill("solid", fgColor=RED)
        ws_d["A1"].alignment = ctr
        ws_d.row_dimensions[1].height = 26
        hdr_row(ws_d, 2, list(detail_df.columns))
        for ri, (_, row) in enumerate(detail_df.iterrows(), 3):
            for ci, v in enumerate(row.values, 1):
                c = ws_d.cell(row=ri, column=ci, value=v)
                c.font=body_font; c.border=bdr; c.alignment=lft
                if ri % 2 == 0: c.fill = PatternFill("solid", fgColor="FFF5F5")
        for ci in range(1, nc+1):
            mx = max(len(str(detail_df.columns[ci-1])),
                     max((len(str(v)) for v in detail_df.iloc[:, ci-1].astype(str)), default=0))
            ws_d.column_dimensions[get_column_letter(ci)].width = min(max(mx+3, 12), 42)

    # helper to write raw data sheet
    def write_raw(ws_r, df):
        ws_r.sheet_view.showGridLines = False
        if df.empty: return
        hdr_row(ws_r, 1, list(df.columns))
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, v in enumerate(row.values, 1):
                c = ws_r.cell(row=ri, column=ci, value=v)
                c.font=body_font; c.border=bdr; c.alignment=lft
                if ri % 2 == 0: c.fill = lgry_fill
        for ci in range(1, len(df.columns)+1):
            ws_r.column_dimensions[get_column_letter(ci)].width = 18

    write_summary(wb.create_sheet("MLoS QC Summary"),       "MLoS Table — QC Check Summary",          mlos_checks)
    write_detail (wb.create_sheet("MLoS Issue Detail"),      "MLoS Table — Detailed Issue Rows",        mlos_detail)
    write_summary(wb.create_sheet("Takeoffpoint QC Summary"),"Takeoffpoint Table — QC Check Summary",   tp_checks)
    write_detail (wb.create_sheet("Takeoffpoint Issue Detail"),"Takeoffpoint Table — Detailed Issues",  tp_detail)
    write_raw    (wb.create_sheet("Raw MLoS Data"),          mlos_df)
    write_raw    (wb.create_sheet("Raw Takeoffpoint Data"),  takeoff_df)

    wb.save(out)
    return out.getvalue()


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📁 Upload File")
    uploaded = st.file_uploader(
        "Upload SQLite, CSV, or Excel file",
        type=["sqlite", "db", "sqlite3", "csv", "xlsx", "xls"],
        key="sqlite_upload", label_visibility="collapsed",
    )
    st.caption("Supported: `.sqlite` · `.csv` · `.xlsx` · `.xls`")
    st.markdown("---")
    st.markdown("#### 📋 QC Rules Reference")
    with st.expander("🔎 Schema Alignment Rules", expanded=False):
        st.markdown("""
| Rule | Check |
|------|-------|
| S1 | All 41 required MLoS columns present |
| S2 | All 4 required Takeoffpoint columns present |

**MLoS required columns:**
`state_code`, `state_name`, `lga_code`, `lga_name`, `ward_name`, `ward_code`, `takeoffpoint`, `takeoffpoint_code`, `settlement_name`, `primarysettlement_name`, `alternate_name`, `latitude`, `longitude`, `security_compromised`, `accessibility_status`, `reasons_for_inaccessibility`, `habitational_status`, `set_population`, `set_target`, `number_of_houses`, `noncompliant_household`, `team_code`, `day_of_activity`, `urban`, `rural`, `highrisk`, `slums`, `densely_populated`, `hard2reach`, `border`, `normadic`, `scattered`, `riverine`, `fulani`, `timestamp`, `source`, `last_updated`, `editor`, `globalid`, `fc_globalid`, `settlementarea_globalid`

**Takeoffpoint required columns:**
`name`, `code`, `wardcode`, `globalid`
        """)
    with st.expander("🏘️ MLoS Table Rules", expanded=False):
        st.markdown("""
| Rule | Check |
|------|-------|
| 2 | takeoffpoint == takeoffpoint.name |
| 3 | takeoffpoint_code == takeoffpoint.code |
| 4 | ward_code == takeoffpoint.wardcode |
| 5 | Required fields not null |
| 6 | security_compromised = Y/N |
| 7 | accessibility_status valid |
| 8 | Partial/Inaccessible requires reason |
| 9 | habitational_status valid |
| 10 | set_target & houses ≤ set_population |
| 12 | day_of_activity valid code |
| 13 | urban/rural/scattered = Y/N (no conflict) |
| 14 | Profile flags = Y/N/NA |
| 15 | source starts with MLoS |
| 16 | editor = firstname.surname (lowercase) |
| 17 | globalid = valid UUID |
        """)
    with st.expander("📍 Takeoffpoint Rules", expanded=False):
        st.markdown("""
| Rule | Check |
|------|-------|
| TP2 | name == mlos.takeoffpoint |
| TP3 | code == mlos.takeoffpoint_code |
| TP4 | wardcode == mlos.ward_code |
| TP5 | globalid = valid UUID |
        """)
    with st.expander("🗺️ Boundary Check Rules", expanded=False):
        st.markdown("""
| Rule | Check |
|------|-------|
| B1 | ward_code exists in admin ward boundary reference (9,410 wards) |
| B2 | latitude/longitude falls within the bounding box of the declared ward_code |
        """)
    st.markdown("---")
    st.caption("eHealth Africa · MLOS QC Tool · v1.1")


# ─── HEADER ───────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <h1>🗺️ MLOS CHECKOUT QC</h1>
  <p>Master List of Settlements — Quality Control Dashboard &nbsp;|&nbsp; Upload a SQLite checkout file to begin</p>
</div>
""", unsafe_allow_html=True)

if not uploaded:
    st.info("👈 **Upload a file** (`.sqlite`, `.csv`, `.xlsx`, or `.xls`) using the sidebar to run QC checks.")
    with st.expander("ℹ️ About this tool", expanded=True):
        st.markdown("""
This tool runs automated Quality Control on MLOS checkout files across **4 QC layers**.

**Supported formats:**
- **SQLite** (`.sqlite`, `.db`) — reads views directly
- **Excel** (`.xlsx`, `.xls`) — Sheet 1 = MLoS, Sheet 2 = Takeoffpoint
- **CSV** (`.csv`) — MLoS data only (takeoffpoint cross-checks skipped)

**QC Layers:**
- 🔎 **Schema Alignment** (S1–S2) — verifies all required columns are present
- 🏘️ **MLoS Rules** (2–17) — 15+ data integrity checks
- 📍 **Takeoffpoint Rules** (TP2–TP5) — 4 cross-table checks
- 🗺️ **Boundary Checks** (B1–B2) — ward code and coordinate validation against admin boundary reference

**Outputs:**
- Pass Rate % and Fail Rate % on the dashboard
- Per-rule issue drilldown with downloadable Excel reports
- Full QC report (Excel) and email summary to the data team
        """)
    st.stop()

# ─── LOAD DATA ────────────────────────────────────────────────────────────────────
load_progress = st.progress(0)
with st.spinner("Loading file…"):
    try:
        mlos_df, takeoff_df = get_uploaded_data(uploaded, progress=load_progress)
    except Exception as e:
        st.error(f"❌ Failed to load file: {e}")
        st.stop()

filename = uploaded.name

# ─── RUN QC ───────────────────────────────────────────────────────────────────────
qc_cache_key = (filename, getattr(uploaded, "size", None))
boundary_ref, boundary_bbox = load_boundary_refs()

if st.session_state.get("qc_cache_key") == qc_cache_key:
    schema_checks    = st.session_state["schema_checks"]
    schema_detail    = st.session_state["schema_detail"]
    mlos_checks      = st.session_state["mlos_checks"]
    mlos_detail      = st.session_state["mlos_detail"]
    tp_checks        = st.session_state["tp_checks"]
    tp_detail        = st.session_state["tp_detail"]
    boundary_checks  = st.session_state["boundary_checks"]
    boundary_detail  = st.session_state["boundary_detail"]
    load_progress.progress(100)
else:
    qc_bar = st.progress(0, text="⏳ Starting QC checks…")
    with st.status("Running QC checks…", expanded=True) as _qc_status:

        # ── Step 1: Schema Alignment ──────────────────────────────────────
        qc_bar.progress(5, text="🔎 Step 1 / 4 — Schema Alignment…")
        st.write("🔎 Step 1 / 4 — Schema Alignment")
        schema_checks, schema_detail = run_schema_qc(mlos_df, takeoff_df)
        schema_fail = sum(1 for c in schema_checks if "FAIL" in c["Status"])
        st.write(f"   {'❌' if schema_fail else '✅'} Schema: "
                 f"{len(schema_checks) - schema_fail}/{len(schema_checks)} checks passed")
        qc_bar.progress(25, text="✅ Step 1 / 4 — Schema Alignment complete")

        # ── Step 2: MLoS Rules ────────────────────────────────────────────
        qc_bar.progress(26, text="🏘️ Step 2 / 4 — MLoS Rules…")
        st.write("🏘️ Step 2 / 4 — MLoS Rules")
        mlos_checks, mlos_detail = run_mlos_qc(mlos_df, takeoff_df)
        mlos_qc_fail = sum(1 for c in mlos_checks if "FAIL" in c["Status"])
        st.write(f"   {'❌' if mlos_qc_fail else '✅'} MLoS: "
                 f"{len(mlos_checks) - mlos_qc_fail}/{len(mlos_checks)} checks passed")
        qc_bar.progress(50, text="✅ Step 2 / 4 — MLoS Rules complete")

        # ── Step 3: Takeoffpoint Rules ────────────────────────────────────
        qc_bar.progress(51, text="📍 Step 3 / 4 — Takeoffpoint Rules…")
        st.write("📍 Step 3 / 4 — Takeoffpoint Rules")
        if takeoff_df.empty:
            tp_checks, tp_detail = [], pd.DataFrame()
            st.write("   ⚠️ Takeoffpoint data not available — skipped")
        else:
            tp_checks, tp_detail = run_takeoff_qc(takeoff_df, mlos_df)
            tp_qc_fail = sum(1 for c in tp_checks if "FAIL" in c["Status"])
            st.write(f"   {'❌' if tp_qc_fail else '✅'} Takeoffpoint: "
                     f"{len(tp_checks) - tp_qc_fail}/{len(tp_checks)} checks passed")
        qc_bar.progress(75, text="✅ Step 3 / 4 — Takeoffpoint Rules complete")

        # ── Step 4: Boundary Checks ───────────────────────────────────────
        qc_bar.progress(76, text="🗺️ Step 4 / 4 — Boundary Checks…")
        st.write("🗺️ Step 4 / 4 — Boundary Checks")
        boundary_checks, boundary_detail = run_ward_boundary_qc(mlos_df, boundary_ref, boundary_bbox)
        b_fail = sum(1 for c in boundary_checks if "FAIL" in c["Status"])
        st.write(f"   {'❌' if b_fail else '✅'} Boundary: "
                 f"{len(boundary_checks) - b_fail}/{len(boundary_checks)} checks passed")
        qc_bar.progress(100, text="✅ All QC checks complete!")

        load_progress.progress(100)
        _qc_status.update(label="QC checks complete!", state="complete", expanded=False)
    st.session_state["qc_cache_key"]   = qc_cache_key
    st.session_state["schema_checks"]  = schema_checks
    st.session_state["schema_detail"]  = schema_detail
    st.session_state["mlos_checks"]    = mlos_checks
    st.session_state["mlos_detail"]    = mlos_detail
    st.session_state["tp_checks"]      = tp_checks
    st.session_state["tp_detail"]      = tp_detail
    st.session_state["boundary_checks"]= boundary_checks
    st.session_state["boundary_detail"]= boundary_detail

all_checks          = schema_checks + mlos_checks + tp_checks + boundary_checks
n_fail              = sum(1 for c in all_checks if "FAIL" in c["Status"])
n_pass              = len(all_checks) - n_fail
pct_pass            = f"{n_pass / len(all_checks) * 100:.1f}%" if all_checks else "0%"
pct_fail            = f"{n_fail / len(all_checks) * 100:.1f}%" if all_checks else "0%"
mlos_fail_rows      = len(mlos_detail)
tp_fail_rows        = len(tp_detail)
boundary_fail_rows  = len(boundary_detail)

# ─── FILE INFO + METRICS ──────────────────────────────────────────────────────────
st.success(f"✅ Loaded **{filename}** — MLoS: **{len(mlos_df):,} rows** | Takeoffpoint: **{len(takeoff_df):,} rows**")

c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
c1.metric("📄 MLoS Rows",      f"{len(mlos_df):,}")
c2.metric("📍 TP Rows",         f"{len(takeoff_df):,}")
c3.metric("🔍 Checks Run",      f"{len(all_checks)}")
c4.metric("✅ Passing",         f"{n_pass}")
c5.metric("❌ Failing",         f"{n_fail}")
c6.metric("⚠️ Issue Rows",     f"{mlos_fail_rows + tp_fail_rows + boundary_fail_rows:,}")
c7.metric("📈 Pass Rate",       pct_pass)
c8.metric("📉 Fail Rate",       pct_fail)

st.markdown("---")

# ─── VERDICT BANNER ───────────────────────────────────────────────────────────────
if n_fail == 0:
    st.markdown('<div class="banner-pass">✅ ALL QC CHECKS PASSED — This file is clean and ready for submission.</div>',
                unsafe_allow_html=True)
else:
    st.markdown(
        f'<div class="banner-fail">❌ {n_fail} QC CHECK(S) FAILING — '
        f'{mlos_fail_rows} MLoS rows and {tp_fail_rows} Takeoffpoint rows have issues. '
        f'Review details below before submission.</div>',
        unsafe_allow_html=True)

# ─── TABS ─────────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📊 QC Summary",
    "🔧 Auto Correct",
    "🏘️ MLoS Issues",
    "📍 Takeoffpoint Issues",
    "🗺️ Boundary Issues",
    "🔍 Raw Data",
    "📄 Generate Report",
])

# ── Tab 1: QC Summary ─────────────────────────────────────────────────────────────
with tab1:
    def colour_rows(row):
        if "FAIL" in str(row.get("Status","")):
            return ["background-color:#fff1f2; color:#be123c"] * len(row)
        return ["background-color:#f0fdf4; color:#15803d"] * len(row)

    # ── Schema Alignment ─────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">🔎 Schema Alignment — QC Results</div>', unsafe_allow_html=True)
    if schema_checks:
        df_s = pd.DataFrame(schema_checks)[["Status","Rule#","QC Check","Description","Failing Rows","Total Rows","Fail %"]]
        df_s = df_s.rename(columns={"Failing Rows": "Missing Columns", "Total Rows": "Expected Columns"})
        st.dataframe(df_s.style.apply(colour_rows, axis=1), use_container_width=True, hide_index=True)
    if not schema_detail.empty:
        with st.expander("🔍 View Missing Column Details", expanded=False):
            st.dataframe(schema_detail.drop(columns=["Rule#","Rule"], errors="ignore"),
                         use_container_width=True, hide_index=True)
        st.download_button(
            label="⬇️ Download Schema Error Report (.xlsx)",
            data=build_schema_report_xlsx(schema_detail),
            file_name=filename.rsplit(".",1)[0] + "_schema_errors.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="schema_dl_tab1",
        )
    else:
        st.success("✅ Schema aligned — all required columns are present.")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── MLoS QC ──────────────────────────────────────────────────────────────────
    st.markdown('<div class="sec-title">🏘️ MLoS Table — QC Results</div>', unsafe_allow_html=True)
    if mlos_checks:
        df_m = pd.DataFrame(mlos_checks)[["Status","Rule#","QC Check","Description","Failing Rows","Total Rows","Fail %"]]
        st.dataframe(df_m.style.apply(colour_rows, axis=1), use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-title">📍 Takeoffpoint Table — QC Results</div>', unsafe_allow_html=True)
    if tp_checks:
        df_t = pd.DataFrame(tp_checks)[["Status","Rule#","QC Check","Description","Failing Rows","Total Rows","Fail %"]]
        st.dataframe(df_t.style.apply(colour_rows, axis=1), use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sec-title">🗺️ Boundary Checks — QC Results</div>', unsafe_allow_html=True)
    if boundary_checks:
        df_b = pd.DataFrame(boundary_checks)[["Status","Rule#","QC Check","Description","Failing Rows","Total Rows","Fail %"]]
        st.dataframe(df_b.style.apply(colour_rows, axis=1), use_container_width=True, hide_index=True)

# ── Tab 2: Auto Correct ───────────────────────────────────────────────────────────
with tab2:
    st.markdown('<div class="sec-title">🔧 Auto Correct — MLoS Data Fixes</div>', unsafe_allow_html=True)
    st.caption(
        "The following corrections are applied automatically to the uploaded MLoS data:\n\n"
        "1. **Flag columns** (`highrisk`, `slums`, `densely_populated`, `hard2reach`, `border`, "
        "`normadic`, `scattered`, `riverine`, `fulani`) — NULL values filled with `NA`\n"
        "2. **Reason for Inaccessibility** — filled with `NA` where `accessibility_status` is "
        "`Fully Accessible` and the reason is NULL\n"
        "3. **GlobalID** — leading/trailing braces `{}` stripped; any still-invalid UUID is "
        "replaced with a freshly generated UUID"
    )
    st.markdown("---")

    corrected_df, correction_log = auto_correct_mlos(mlos_df)

    if not correction_log:
        st.success("✅ No corrections needed — all checked fields are already valid.")
    else:
        total_fixes = sum(r["Rows Fixed"] for r in correction_log)
        st.info(f"🔧 **{len(correction_log)} correction(s) applied** across **{total_fixes:,} row(s)**.")
        log_df = pd.DataFrame(correction_log)
        st.dataframe(log_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**📥 Download Corrected MLoS File**")
        st.caption("The corrected data is exported as an Excel file ready for re-upload or submission.")
        corrected_bytes = build_corrected_excel(corrected_df)
        corrected_name  = filename.rsplit(".", 1)[0] + "_corrected.xlsx"
        st.download_button(
            label     = "⬇️ Download Corrected MLoS (Excel)",
            data      = corrected_bytes,
            file_name = corrected_name,
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type      = "primary",
        )


# ── Tab 3: MLoS Issues ────────────────────────────────────────────────────────────
with tab3:
    if mlos_detail.empty:
        st.success("✅ No issues found in the MLoS table — all checks passed!")
    else:
        n_failing_checks = sum(1 for c in mlos_checks if "FAIL" in c["Status"])
        st.error(f"❌ **{mlos_fail_rows:,} issue row(s)** across **{n_failing_checks} failing check(s)**")

        for check in mlos_checks:
            if "FAIL" not in check["Status"]: continue
            rn     = check["Rule#"]
            subset = mlos_detail[mlos_detail["Rule#"] == rn].drop(columns=["Rule#","Rule"], errors="ignore")
            n      = len(subset)
            with st.expander(f"❌  Rule {rn} — {check['QC Check']}  ({n:,} row{'s' if n!=1 else ''})", expanded=False):
                st.caption(f"📌 {check['Description']}")
                st.dataframe(subset, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**📋 MLoS Issue Rows — Longitudinal View**")
        st.caption("One row per settlement. Each QC rule appears as a column: **Yes** = rule error present, **No** = no error.")

        long_df = make_longitudinal_df(mlos_df, mlos_checks, mlos_detail)
        if not long_df.empty:
            rule_flag_cols = [c for c in long_df.columns if str(c).startswith("Rule_")]
            # Convert booleans to emoji strings for visual clarity — avoids Styler compatibility issues
            display_df = long_df.copy()
            for col in rule_flag_cols:
                display_df[col] = display_df[col].map(lambda v: "Yes" if v is True else ("No" if v is False else v))
            st.dataframe(display_df, use_container_width=True, hide_index=True, height=350)

        long_xlsx = build_longitudinal_mlos(mlos_df, mlos_checks, mlos_detail)
        st.download_button(
            "⬇️ Download MLoS Issues — Longitudinal (Excel)",
            data=long_xlsx,
            file_name=filename.rsplit(".", 1)[0] + "_mlos_issues_longitudinal.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ── Tab 4: Takeoffpoint Issues ────────────────────────────────────────────────────
with tab4:
    if tp_detail.empty:
        st.success("✅ No issues found in the Takeoffpoint table — all checks passed!")
    else:
        n_failing_tp = sum(1 for c in tp_checks if "FAIL" in c["Status"])
        st.error(f"❌ **{tp_fail_rows:,} issue row(s)** across **{n_failing_tp} failing check(s)**")

        for check in tp_checks:
            if "FAIL" not in check["Status"]: continue
            rn     = check["Rule#"]
            subset = tp_detail[tp_detail["Rule#"] == rn].drop(columns=["Rule#","Rule"], errors="ignore")
            n      = len(subset)
            with st.expander(f"❌  Rule {rn} — {check['QC Check']}  ({n:,} row{'s' if n!=1 else ''})", expanded=False):
                st.caption(f"📌 {check['Description']}")
                st.dataframe(subset, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**📋 All Takeoffpoint Issue Rows (combined)**")
        st.dataframe(tp_detail, use_container_width=True, hide_index=True, height=350)

        buf2 = BytesIO()
        with pd.ExcelWriter(buf2, engine="openpyxl") as xw:
            tp_detail.to_excel(xw, sheet_name="All Issues", index=False)
            for check in tp_checks:
                if "FAIL" not in check["Status"]: continue
                rn  = check["Rule#"]
                sub = tp_detail[tp_detail["Rule#"] == rn].drop(columns=["Rule#","Rule"], errors="ignore")
                sub.to_excel(xw, sheet_name=f"Rule {rn}"[:31], index=False)
        st.download_button("⬇️ Download Takeoffpoint Issues (Excel)", data=buf2.getvalue(),
                           file_name=filename.replace(".sqlite","")+"_tp_issues.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Tab 5: Boundary Issues ────────────────────────────────────────────────────────
with tab5:
    if boundary_detail.empty:
        st.success("✅ All ward codes match the boundary reference and all coordinates fall within their declared ward boundaries.")
    else:
        n_b_fail = sum(1 for c in boundary_checks if "FAIL" in c["Status"])
        st.error(f"❌ **{boundary_fail_rows:,} issue row(s)** across **{n_b_fail} boundary check(s)**")

        for check in boundary_checks:
            if "FAIL" not in check["Status"]: continue
            rn     = check["Rule#"]
            subset = boundary_detail[boundary_detail["Rule#"] == rn].drop(columns=["Rule#","Rule"], errors="ignore")
            n      = len(subset)
            with st.expander(f"❌  Rule {rn} — {check['QC Check']}  ({n:,} row{'s' if n!=1 else ''})", expanded=False):
                st.caption(f"📌 {check['Description']}")
                st.dataframe(subset, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**📋 All Boundary Issue Rows (combined)**")
        st.dataframe(boundary_detail, use_container_width=True, hide_index=True, height=350)

        buf_b = BytesIO()
        with pd.ExcelWriter(buf_b, engine="openpyxl") as xw:
            boundary_detail.to_excel(xw, sheet_name="All Boundary Issues", index=False)
            for check in boundary_checks:
                if "FAIL" not in check["Status"]: continue
                rn  = check["Rule#"]
                sub = boundary_detail[boundary_detail["Rule#"] == rn].drop(columns=["Rule#","Rule"], errors="ignore")
                sub.to_excel(xw, sheet_name=f"Rule {rn}"[:31], index=False)
        st.download_button("⬇️ Download Boundary Issues (Excel)", data=buf_b.getvalue(),
                           file_name=filename.rsplit(".",1)[0] + "_boundary_issues.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Tab 6: Raw Data ───────────────────────────────────────────────────────────────
with tab6:
    rt1, rt2 = st.tabs(["🏘️ MLoS View", "📍 Takeoffpoint View"])
    with rt1:
        st.markdown(f"**{len(mlos_df):,} rows × {len(mlos_df.columns)} columns**")
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            state_opts = ["All"] + sorted(mlos_df["state_name"].dropna().unique().tolist()) \
                         if "state_name" in mlos_df.columns else ["All"]
            sel_state = st.selectbox("State", state_opts, key="sel_state")
        with fc2:
            if sel_state != "All" and "lga_name" in mlos_df.columns:
                lga_opts = ["All"] + sorted(mlos_df[mlos_df["state_name"]==sel_state]["lga_name"].dropna().unique().tolist())
            elif "lga_name" in mlos_df.columns:
                lga_opts = ["All"] + sorted(mlos_df["lga_name"].dropna().unique().tolist())
            else:
                lga_opts = ["All"]
            sel_lga = st.selectbox("LGA", lga_opts, key="sel_lga")
        with fc3:
            search = st.text_input("Search settlement", placeholder="Type to filter…", key="srch")

        fdf = mlos_df.copy()
        if sel_state != "All" and "state_name" in fdf.columns:
            fdf = fdf[fdf["state_name"] == sel_state]
        if sel_lga != "All" and "lga_name" in fdf.columns:
            fdf = fdf[fdf["lga_name"] == sel_lga]
        if search and "settlement_name" in fdf.columns:
            fdf = fdf[fdf["settlement_name"].astype(str).str.contains(search, case=False, na=False)]
        st.dataframe(fdf, use_container_width=True, hide_index=True, height=420)

    with rt2:
        st.markdown(f"**{len(takeoff_df):,} rows × {len(takeoff_df.columns)} columns**")
        st.dataframe(takeoff_df, use_container_width=True, hide_index=True, height=420)

# ── Tab 7: Generate Report ────────────────────────────────────────────────────────
with tab7:
    st.markdown('<div class="sec-title">📄 QC Report — Summary &amp; Download</div>', unsafe_allow_html=True)

    # Verdict
    if n_fail == 0:
        st.markdown(
            '<div class="report-verdict-pass">✅ &nbsp; ALL CHECKS PASSED — File is CLEAN and ready for submission</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div class="report-verdict-fail">❌ &nbsp; {n_fail} CHECK(S) FAILING — '
            f'{mlos_fail_rows + tp_fail_rows} total issue rows. Fix before submission.</div>',
            unsafe_allow_html=True)

    # Metadata card
    gen_time      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mlos_n_fail   = sum(1 for c in mlos_checks if "FAIL" in c["Status"])
    mlos_n_pass   = len(mlos_checks) - mlos_n_fail
    tp_n_fail     = sum(1 for c in tp_checks   if "FAIL" in c["Status"])
    tp_n_pass     = len(tp_checks)   - tp_n_fail

    def rrow(lbl, val, cls="val"):
        return (f'<div class="report-row">'
                f'<span class="lbl">{lbl}</span>'
                f'<span class="{cls}">{val}</span>'
                f'</div>')

    html = "".join([
        rrow("📁 File",                    filename),
        rrow("🕐 Generated At",            gen_time),
        rrow("━" * 36,                     "", "val-div"),
        rrow("MLoS Rows",                  f"{len(mlos_df):,}"),
        rrow("Takeoffpoint Rows",          f"{len(takeoff_df):,}"),
        rrow("━" * 36,                     "", "val-div"),
        rrow("MLoS Checks Run",            str(len(mlos_checks))),
        rrow("MLoS Checks ✅ Passing",     str(mlos_n_pass),  "val-pass"),
        rrow("MLoS Checks ❌ Failing",     str(mlos_n_fail),  "val-fail" if mlos_n_fail else "val-pass"),
        rrow("MLoS Issue Rows",            f"{mlos_fail_rows:,}",
                                            "val-fail" if mlos_fail_rows else "val-pass"),
        rrow("━" * 36,                     "", "val-div"),
        rrow("Takeoffpoint Checks Run",    str(len(tp_checks))),
        rrow("Takeoffpoint Checks ✅ Passing", str(tp_n_pass),  "val-pass"),
        rrow("Takeoffpoint Checks ❌ Failing", str(tp_n_fail),  "val-fail" if tp_n_fail else "val-pass"),
        rrow("Takeoffpoint Issue Rows",    f"{tp_fail_rows:,}",
                                            "val-fail" if tp_fail_rows else "val-pass"),
    ])
    st.markdown(f'<div class="report-card">{html}</div>', unsafe_allow_html=True)

    # Side-by-side check breakdown
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**🏘️ MLoS Check Breakdown**")
        tbl_m = pd.DataFrame([{"Status": c["Status"], "Rule#": c["Rule#"],
                                "QC Check": c["QC Check"], "Failing": c["Failing Rows"],
                                "Fail %": c["Fail %"]} for c in mlos_checks])
        def hl(row):
            return (["color:#be123c; font-weight:700"] + [""]*4 if "FAIL" in str(row.get("Status",""))
                    else ["color:#15803d; font-weight:700"] + [""]*4)
        st.dataframe(tbl_m.style.apply(hl, axis=1), use_container_width=True, hide_index=True)

    with col_b:
        st.markdown("**📍 Takeoffpoint Check Breakdown**")
        tbl_t = pd.DataFrame([{"Status": c["Status"], "Rule#": c["Rule#"],
                                "QC Check": c["QC Check"], "Failing": c["Failing Rows"],
                                "Fail %": c["Fail %"]} for c in tp_checks])
        st.dataframe(tbl_t.style.apply(hl, axis=1), use_container_width=True, hide_index=True)

    st.markdown("---")

    # Download
    st.markdown("**📥 Download Full QC Report**")
    st.caption("7-sheet Excel report: Cover Page · MLoS QC Summary · MLoS Issue Detail · "
               "Takeoffpoint QC Summary · Takeoffpoint Issue Detail · Raw MLoS Data · Raw Takeoffpoint Data")

    with st.spinner("Building Excel report…"):
        report_bytes = build_excel_report(
            filename, mlos_checks, mlos_detail,
            tp_checks, tp_detail, mlos_df, takeoff_df)

    report_file = filename.replace(".sqlite","").replace(".db","") + "_QC_Report.xlsx"
    st.download_button(
        label   = "⬇️  Download Full QC Report (Excel)",
        data    = report_bytes,
        file_name = report_file,
        mime    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type    = "primary",
    )

    if n_fail == 0:
        st.success("✅ File is CLEAN — all checks passed.")
    else:
        st.warning(f"⚠️ {n_fail} check(s) failing with "
                   f"{mlos_fail_rows + tp_fail_rows + boundary_fail_rows} issue rows. "
                   f"Fix and re-upload before submission.")

    st.markdown("---")
    st.markdown("**📧 Send QC Summary by Email**")
    st.caption("Sends the QC summary to the data team. SMTP credentials must be configured in Streamlit secrets.")
    if st.button("📤 Send QC Email", type="primary", key="send_email_btn"):
        try:
            send_qc_email(filename, all_checks, mlos_fail_rows,
                          tp_fail_rows, schema_detail, boundary_fail_rows)
            st.success("✅ Email sent successfully to the data team.")
        except Exception as e:
            st.error(f"❌ Failed to send email: {e}"
                     "\n\nEnsure `smtp_host`, `smtp_port`, `smtp_user`, and `smtp_pass` "
                     "are set in your Streamlit secrets.")
